#!/usr/bin/env python3
"""monitoring/report/append-daily-sheet.selftest.py

Offline self-test for append-daily-sheet.py. Makes NO network calls and
touches ONLY temp-directory paths (a synthetic report JSON + a temp .xlsx) —
never a real monitoring/reports/*.json or the real accumulating workbook.

Runs the appender TWICE for the SAME date into the SAME temp workbook (the
"same-day re-run" case) and asserts, via openpyxl, that the result is
idempotent:
  1. Exactly ONE sheet is named the report's date (no duplicate, e.g. no
     "2026-07-29" AND "2026-07-29 (2)").
  2. That sheet contains the expected tool rows (status/tool/domain/severity
     /ms/error), sorted worst-first (FAIL before SLOW before PASS).
  3. The `Trends` sheet has exactly ONE row for that date (not two, even
     though the appender ran twice) and is the LAST sheet in the workbook.
  4. The `_meta` sheet is hidden and also has exactly one row for the date.

Run after ANY change to append-daily-sheet.py:
    python3 monitoring/report/append-daily-sheet.selftest.py

Exit code: 0 on all assertions passing, 1 otherwise (also prints PASS/FAIL).
"""

from __future__ import annotations

import json
import os
import shutil
import subprocess
import sys
import tempfile

from openpyxl import load_workbook

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
APPENDER = os.path.join(SCRIPT_DIR, "append-daily-sheet.py")

TEST_DATE = "2026-07-29"


def make_synthetic_report() -> dict:
    return {
        "startedAt": f"{TEST_DATE}T02:57:00.000Z",
        "finishedAt": f"{TEST_DATE}T03:09:12.345Z",
        "mcpUrl": "https://mcp.novada.com/mcp",
        "liveToolCount": 3,
        "exitCode": 0,
        "summary": {
            "maxSeverity": "P1",
            "maxOursSeverity": None,
            "byStatus": {"PASS": 1, "SLOW": 1, "FAIL": 1},
            "bySeverity": {"P1": 1},
            "oursCount": 0,
            "backendCount": 1,
            "missingTools": [],
        },
        "results": [
            {
                "name": "novada_setup",
                "platform": "-",
                "operation": "-",
                "status": "PASS",
                "domain": "-",
                "severity": None,
                "timeMs": 365,
                "taskId": None,
                "error": None,
                "advice": "-",
            },
            {
                "name": "novada_research",
                "platform": "-",
                "operation": "-",
                "status": "SLOW",
                "domain": "-",
                "severity": None,
                "timeMs": 50989,
                "taskId": None,
                "error": None,
                "advice": "-",
            },
            {
                "name": "novada_scrape_amazon",
                "platform": "amazon.com",
                "operation": "amazon_product_keywords",
                "status": "FAIL",
                "domain": "③-backend",
                "severity": "P1",
                "timeMs": 1200,
                "taskId": "task-123",
                "error": "Scraper error (code 50004): context deadline exceeded",
                "advice": "escalate to backend",
            },
        ],
    }


failure_count = 0


def expect(condition: bool, message: str) -> None:
    global failure_count
    if condition:
        print(f"  ✓ {message}")
    else:
        failure_count += 1
        print(f"  ✗ FAIL: {message}")


def main() -> int:
    tmp_dir = tempfile.mkdtemp(prefix="append-daily-sheet-selftest-")
    try:
        report_path = os.path.join(tmp_dir, f"full-{TEST_DATE}T02-57-00-000Z.json")
        xlsx_path = os.path.join(tmp_dir, "daily-monitor-selftest.xlsx")

        with open(report_path, "w", encoding="utf-8") as f:
            json.dump(make_synthetic_report(), f)

        print(f"[selftest] running appender TWICE for the same date ({TEST_DATE}) into {xlsx_path} ...")
        for i in (1, 2):
            result = subprocess.run(
                [sys.executable, APPENDER, report_path, xlsx_path],
                capture_output=True,
                text=True,
            )
            print(f"  run {i}: exit={result.returncode} stdout={result.stdout.strip()!r}")
            expect(result.returncode == 0, f"run {i} of the appender exits 0 (stderr: {result.stderr.strip()})")

        wb = load_workbook(xlsx_path)

        # ── Assertion 1: exactly one sheet named the date (idempotent replace) ──
        date_sheets = [name for name in wb.sheetnames if name == TEST_DATE]
        expect(len(date_sheets) == 1, f"exactly one sheet named {TEST_DATE} (got {date_sheets!r} in {wb.sheetnames!r})")

        # ── Assertion 2: expected tool rows are present, worst-first sorted ──
        ws = wb[TEST_DATE]
        rows = list(ws.iter_rows(values_only=True))
        row_texts = [r for r in rows if r and r[0] in ("✅", "⚠️", "❌")]
        expect(len(row_texts) == 3, f"3 tool rows present in the dated sheet (got {len(row_texts)}: {row_texts})")

        names_in_order = [r[1] for r in row_texts]
        expect(
            names_in_order == ["novada_scrape_amazon", "novada_research", "novada_setup"],
            f"rows sorted FAIL, then SLOW, then PASS (got order: {names_in_order})",
        )

        fail_row = next((r for r in row_texts if r[1] == "novada_scrape_amazon"), None)
        expect(fail_row is not None and fail_row[0] == "❌", "the FAIL row shows the ❌ status icon")
        expect(
            fail_row is not None and fail_row[6] and "50004" in str(fail_row[6]),
            f"the FAIL row's error column contains the original error text (got: {fail_row[6] if fail_row else None!r})",
        )

        # ── Assertion 3: Trends has exactly one row for the date, and is last ──
        expect(wb.sheetnames[-1] == "Trends", f"'Trends' is the LAST sheet (got sheet order: {wb.sheetnames!r})")
        ws_trends = wb["Trends"]
        trends_rows = [r for r in ws_trends.iter_rows(min_row=2, values_only=True) if r and r[0] == TEST_DATE]
        expect(
            len(trends_rows) == 1,
            f"Trends has exactly one row for {TEST_DATE} even after 2 appender runs (got {len(trends_rows)})",
        )
        if trends_rows:
            expect(
                trends_rows[0][1:4] == (1, 1, 1),
                f"Trends row has the right PASS/SLOW/FAIL counts (got {trends_rows[0][1:4]}, expected (1, 1, 1))",
            )

        # ── Assertion 4: _meta is hidden and has exactly one row for the date ──
        expect("_meta" in wb.sheetnames, "'_meta' sheet exists")
        ws_meta = wb["_meta"]
        expect(ws_meta.sheet_state == "hidden", f"'_meta' sheet is hidden (got sheet_state={ws_meta.sheet_state!r})")
        meta_rows = [r for r in ws_meta.iter_rows(min_row=2, values_only=True) if r and r[0] == TEST_DATE]
        expect(len(meta_rows) == 1, f"_meta has exactly one row for {TEST_DATE} (got {len(meta_rows)})")

        print("")
        if failure_count > 0:
            print(f"[selftest] FAILED: {failure_count} assertion(s) did not hold.")
            return 1
        print("[selftest] PASS — idempotent replace, sort order, Trends, and _meta all verified.")
        return 0
    finally:
        shutil.rmtree(tmp_dir, ignore_errors=True)


if __name__ == "__main__":
    sys.exit(main())
