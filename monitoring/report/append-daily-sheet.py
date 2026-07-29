#!/usr/bin/env python3
"""monitoring/report/append-daily-sheet.py

Appends ONE day of a Layer D `monitoring/reports/full-<ts>.json` probe report
as a new dated sheet inside a single, ACCUMULATING local Excel workbook — "one
Excel file, one sheet per day" — so the owner can open ONE file and see every
day's tool-health snapshot without hunting through monitoring/reports/.

Usage:
    python3 monitoring/report/append-daily-sheet.py <report-json-path> <excel-path>

Behavior:
  - If <excel-path> doesn't exist yet, a new workbook is created.
  - The date (YYYY-MM-DD) is derived from the report's `startedAt` (UTC,
    falls back to `finishedAt` then today's UTC date if both are missing —
    see the full-tools-probe.mjs FATAL path, which writes `startedAt: null`).
  - A sheet named that date is added at index 0 (newest-first tab order, so
    opening the file lands on today). If a sheet with that name already
    exists (a same-day re-run), it is DELETED and re-added — idempotent,
    never duplicated.
  - A hidden `_meta` sheet stores one row per day (date, PASS, SLOW, FAIL,
    oursCount, backendCount, exitCode). This is the SIMPLEST robust way to
    rebuild the `Trends` sheet on every run: re-parsing formatted title-block
    text back out of each dated sheet would be fragile (text wording could
    drift); a small structured ledger never has to be re-parsed, only
    upserted (by date) and read back. `_meta` is hidden because it is
    plumbing, not something the owner needs to look at.
  - `Trends` (kept as the LAST sheet) is fully REBUILT every run from `_meta`
    — one row per day, newest first.

No network calls, no secrets. Python3 + openpyxl only.
"""

from __future__ import annotations

import json
import os
import sys
from datetime import datetime, timezone
from typing import Any, Optional

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

# ─── Style constants (brand purple header, matches monitoring/report/render-report.py
# and /tmp/gen-explained.py's precedent — see ~/.claude memory
# reference_novada_brand_assets.md) ──────────────────────────────────────────
HEADER_FILL = PatternFill(start_color="271472", end_color="271472", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)
THIN_BORDER = Border(
    left=Side(style="thin", color="D8D8D8"),
    right=Side(style="thin", color="D8D8D8"),
    top=Side(style="thin", color="D8D8D8"),
    bottom=Side(style="thin", color="D8D8D8"),
)
WRAP = Alignment(wrap_text=True, vertical="top")
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)

STATUS_ICON = {"PASS": "✅", "SLOW": "⚠️", "FAIL": "❌", "MISSING": "❌"}
# Fills exactly as specified: red / yellow / green.
STATUS_FILL = {
    "PASS": PatternFill(start_color="E7F6EC", end_color="E7F6EC", fill_type="solid"),
    "SLOW": PatternFill(start_color="FFF6D9", end_color="FFF6D9", fill_type="solid"),
    "FAIL": PatternFill(start_color="FDE7E7", end_color="FDE7E7", fill_type="solid"),
    "MISSING": PatternFill(start_color="FDE7E7", end_color="FDE7E7", fill_type="solid"),
}
# Worst -> best. Anything unrecognized sorts last (rank 9).
STATUS_RANK = {"MISSING": 0, "FAIL": 1, "SLOW": 2, "PASS": 3}

TABLE_HEADERS = ["Status", "Tool", "Domain", "Severity", "ms", "task_id", "error"]
TABLE_COL_WIDTHS = [8, 28, 14, 10, 8, 28, 60]

META_SHEET = "_meta"
META_HEADERS = ["date", "PASS", "SLOW", "FAIL", "oursCount", "backendCount", "exitCode"]
TRENDS_SHEET = "Trends"

ERROR_TRUNCATE_CHARS = 120


def derive_date(report: dict[str, Any]) -> str:
    """YYYY-MM-DD from startedAt, falling back to finishedAt, then today (UTC)."""
    for key in ("startedAt", "finishedAt"):
        value = report.get(key)
        if value:
            return str(value)[:10]
    return datetime.now(timezone.utc).strftime("%Y-%m-%d")


def truncate_error(value: Any) -> str:
    if value is None:
        return "-"
    text = str(value).replace("\n", " ").replace("\r", " ").strip()
    if not text:
        return "-"
    if len(text) > ERROR_TRUNCATE_CHARS:
        return text[:ERROR_TRUNCATE_CHARS] + "…"
    return text


def cell_or_dash(value: Any) -> Any:
    return "-" if value in (None, "") else value


def load_or_create_workbook(path: str) -> tuple[Workbook, Optional[Any]]:
    """Loads the workbook at `path`, or creates a fresh one.

    Returns (workbook, default_sheet_to_remove). For a brand-new workbook,
    `default_sheet_to_remove` is a direct reference to openpyxl's
    auto-created blank "Sheet" — we hold the OBJECT (not its title) so we can
    safely remove it later regardless of naming, once real sheets exist
    (openpyxl refuses to remove the last remaining sheet in a workbook).
    """
    if os.path.exists(path):
        return load_workbook(path), None
    wb = Workbook()
    return wb, wb.worksheets[0]


def get_or_create_meta_sheet(wb: Workbook) -> Worksheet:
    if META_SHEET in wb.sheetnames:
        return wb[META_SHEET]
    ws = wb.create_sheet(META_SHEET)
    ws.append(META_HEADERS)
    ws.sheet_state = "hidden"
    return ws


def upsert_meta_row(ws_meta: Worksheet, date_str: str, counts: dict[str, Any]) -> None:
    """Overwrites the existing row for `date_str`, or appends a new one."""
    row_values = [
        date_str,
        counts.get("PASS", 0),
        counts.get("SLOW", 0),
        counts.get("FAIL", 0),
        counts.get("oursCount", 0),
        counts.get("backendCount", 0),
        counts.get("exitCode", "-"),
    ]
    for row in ws_meta.iter_rows(min_row=2, max_row=ws_meta.max_row):
        if row[0].value == date_str:
            for col_idx, value in enumerate(row_values, start=1):
                ws_meta.cell(row=row[0].row, column=col_idx, value=value)
            return
    ws_meta.append(row_values)


def style_header_row(ws: Worksheet, row_idx: int, num_cols: int) -> None:
    for col_idx in range(1, num_cols + 1):
        cell = ws.cell(row=row_idx, column=col_idx)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.border = THIN_BORDER


def build_day_sheet(wb: Workbook, date_str: str, report: dict[str, Any]) -> tuple[int, int, int]:
    """(Re)builds the dated sheet at index 0. Returns (pass_n, slow_n, fail_n).

    Rows are written with explicit `ws.cell(row=..., column=..., value=...)`
    calls tracked by a local `row` counter, NOT `ws.append([])` — openpyxl's
    `append` only advances `max_row` when at least one cell is actually
    written, so `ws.append([])` (a "blank spacer row") never increments
    `max_row` at all. Relying on `max_row` to detect "have I reached the
    target row yet" therefore spins forever on an empty append. Explicit row
    tracking sidesteps that footgun entirely.
    """
    if date_str in wb.sheetnames:
        del wb[date_str]
    ws = wb.create_sheet(date_str, 0)

    summary = report.get("summary", {}) or {}
    by_status = summary.get("byStatus", {}) or {}
    pass_n = by_status.get("PASS", 0)
    slow_n = by_status.get("SLOW", 0)
    fail_n = by_status.get("FAIL", 0)
    ours_n = summary.get("oursCount", 0)
    backend_n = summary.get("backendCount", 0)
    exit_code = report.get("exitCode", "-")

    num_cols = len(TABLE_HEADERS)
    last_col_letter = get_column_letter(num_cols)

    def write_header_line(row_idx: int, text: str) -> None:
        ws.cell(row=row_idx, column=1, value=text)
        ws.merge_cells(f"A{row_idx}:{last_col_letter}{row_idx}")

    # ── Title / header block ────────────────────────────────────────────
    row = 1
    write_header_line(row, f"Novada MCP Daily Monitor — {date_str}")
    ws.cell(row=row, column=1).font = Font(bold=True, size=13)
    row += 1

    started = cell_or_dash(report.get("startedAt"))
    finished = cell_or_dash(report.get("finishedAt"))
    write_header_line(row, f"Window (UTC): {started} → {finished}")
    row += 1

    write_header_line(row, f"MCP URL: {cell_or_dash(report.get('mcpUrl'))}")
    row += 1

    write_header_line(
        row,
        f"Results: PASS {pass_n} / SLOW {slow_n} / FAIL {fail_n}   |   "
        f"Ours issues: {ours_n}   Backend issues: {backend_n}   |   Probe exit code: {exit_code}",
    )
    row += 1

    missing_tools = summary.get("missingTools") or []
    if missing_tools:
        write_header_line(row, f"Missing tools: {', '.join(missing_tools)}")
        row += 1

    # On the one day this pipeline matters most — a probe that CRASHED before
    # doing any tool calls (full-tools-probe.mjs's fallback writes
    # summary.fatalError with the reason) — surface WHY instead of an unexplained
    # all-zero blank sheet, so the owner doesn't have to dig through raw JSON.
    fatal_error = summary.get("fatalError")
    if fatal_error and not (report.get("results") or []):
        write_header_line(row, f"⚠️ PROBE CRASHED — no tool calls ran: {truncate_error(fatal_error)}")
        ws.cell(row=row, column=1).font = Font(bold=True, color="B00020")
        row += 1

    row += 1  # blank spacer row before the table (deliberately left empty)

    # ── Tool table ───────────────────────────────────────────────────────
    table_header_row = row
    for col_idx, header in enumerate(TABLE_HEADERS, start=1):
        ws.cell(row=table_header_row, column=col_idx, value=header)
    style_header_row(ws, table_header_row, num_cols)

    results = list(report.get("results", []) or [])
    results.sort(key=lambda r: (STATUS_RANK.get(r.get("status"), 9), str(r.get("name", ""))))

    data_row = table_header_row
    for r in results:
        data_row += 1
        status = r.get("status") or "-"
        values = [
            STATUS_ICON.get(status, status),
            cell_or_dash(r.get("name")),
            cell_or_dash(r.get("domain")),
            cell_or_dash(r.get("severity")),
            cell_or_dash(r.get("timeMs")),
            cell_or_dash(r.get("taskId")),
            truncate_error(r.get("error")),
        ]
        fill = STATUS_FILL.get(status)
        for col_idx, value in enumerate(values, start=1):
            cell = ws.cell(row=data_row, column=col_idx, value=value)
            cell.border = THIN_BORDER
            cell.alignment = WRAP
        if fill is not None:
            ws.cell(row=data_row, column=1).fill = fill

    for col_idx, width in enumerate(TABLE_COL_WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.freeze_panes = f"A{table_header_row + 1}"
    ws.auto_filter.ref = f"A{table_header_row}:{last_col_letter}{max(data_row, table_header_row)}"

    return pass_n, slow_n, fail_n


def rebuild_trends_sheet(wb: Workbook, ws_meta: Worksheet) -> None:
    if TRENDS_SHEET in wb.sheetnames:
        del wb[TRENDS_SHEET]
    ws = wb.create_sheet(TRENDS_SHEET)  # appended at the end -> stays LAST.

    headers = ["Date", "PASS", "SLOW", "FAIL", "oursCount", "backendCount", "exitCode"]
    ws.append(headers)
    style_header_row(ws, 1, len(headers))

    meta_rows = [
        [cell.value for cell in row]
        for row in ws_meta.iter_rows(min_row=2, max_row=ws_meta.max_row)
        if row[0].value
    ]
    meta_rows.sort(key=lambda r: str(r[0]), reverse=True)  # ISO dates sort lexically -> newest first.

    for row_values in meta_rows:
        ws.append(row_values)
        row_idx = ws.max_row
        for col_idx in range(1, len(headers) + 1):
            ws.cell(row=row_idx, column=col_idx).border = THIN_BORDER

    for col_idx, width in enumerate([14, 8, 8, 8, 12, 14, 10], start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{ws.max_row}"


def count_day_sheets(wb: Workbook) -> int:
    return len([name for name in wb.sheetnames if name not in (META_SHEET, TRENDS_SHEET)])


def main() -> int:
    if len(sys.argv) != 3:
        print(
            "usage: python3 append-daily-sheet.py <report-json-path> <excel-path>",
            file=sys.stderr,
        )
        return 2

    report_path, excel_path = sys.argv[1], sys.argv[2]

    with open(report_path, "r", encoding="utf-8") as f:
        report = json.load(f)

    date_str = derive_date(report)

    wb, default_sheet = load_or_create_workbook(excel_path)

    pass_n, slow_n, fail_n = build_day_sheet(wb, date_str, report)

    ws_meta = get_or_create_meta_sheet(wb)
    summary = report.get("summary", {}) or {}
    upsert_meta_row(
        ws_meta,
        date_str,
        {
            "PASS": pass_n,
            "SLOW": slow_n,
            "FAIL": fail_n,
            "oursCount": summary.get("oursCount", 0),
            "backendCount": summary.get("backendCount", 0),
            "exitCode": report.get("exitCode", "-"),
        },
    )

    rebuild_trends_sheet(wb, ws_meta)

    # Safe to drop openpyxl's auto-created blank default sheet now — real
    # sheets (the dated sheet, _meta, Trends) already exist, so this never
    # trips openpyxl's "cannot remove the last sheet" guard.
    if default_sheet is not None and default_sheet.title in wb.sheetnames:
        wb.remove(default_sheet)

    os.makedirs(os.path.dirname(os.path.abspath(excel_path)) or ".", exist_ok=True)
    wb.save(excel_path)

    day_sheet_count = count_day_sheets(wb)
    print(
        f"wrote sheet {date_str}: {len(report.get('results', []) or [])} tools "
        f"(PASS {pass_n} / SLOW {slow_n} / FAIL {fail_n}); workbook now has "
        f"{day_sheet_count} day-sheets -> {excel_path}"
    )
    return 0


if __name__ == "__main__":
    sys.exit(main())
