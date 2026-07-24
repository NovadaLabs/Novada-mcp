#!/usr/bin/env python3
"""monitoring/report/render-report.py

Renders the JSON report written by monitoring/smoke/full-tools-probe.mjs
(Layer D — daily full-tools probe) into a severity-sorted .xlsx workbook and a
UTF-8-BOM .csv, both dropped into monitoring/reports/ alongside the source
JSON.

This script reads NO secrets and makes NO network calls — it is a pure
local/CI rendering step. Python + openpyxl only (CI installs openpyxl via
`pip install openpyxl`; it is already installed locally).

Usage:
    python3 monitoring/report/render-report.py [path/to/full-<ts>.json]

If no path is given, the most recently modified monitoring/reports/full-*.json
file is used (the "latest run" — the natural default for a CI step that runs
immediately after monitoring/smoke/full-tools-probe.mjs).

Output (next to the source JSON, same <ts> stem):
    monitoring/reports/full-<ts>.xlsx  — two sheets: 逐工具测试结果, 汇总
    monitoring/reports/full-<ts>.csv   — 逐工具测试结果 only, UTF-8 BOM
"""

from __future__ import annotations

import csv
import glob
import json
import os
import sys
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

MONITORING_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
REPORTS_DIR = os.path.join(MONITORING_DIR, "reports")

# ─── Style constants (matches the brand palette used elsewhere in this repo:
# brand purple #7749F1 -> #271472 — see ~/.claude memory reference_novada_brand_assets.md) ──
HEADER_FILL = PatternFill(start_color="271472", end_color="271472", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)

STATUS_FILLS = {
    "PASS": PatternFill(start_color="D9F2E3", end_color="D9F2E3", fill_type="solid"),  # green
    "SLOW": PatternFill(start_color="FCE8B2", end_color="FCE8B2", fill_type="solid"),  # amber
    "FAIL": PatternFill(start_color="F8D0CE", end_color="F8D0CE", fill_type="solid"),  # red
    "MISSING": PatternFill(start_color="D9534F", end_color="D9534F", fill_type="solid"),  # dark red
}
STATUS_FONTS = {
    "MISSING": Font(color="FFFFFF", bold=True),
}

# Report column order — Chinese headers as specified by the task brief.
COLUMNS: list[tuple[str, str]] = [
    ("状态", "status"),
    ("工具", "name"),
    ("平台/目标", "platform"),
    ("operation", "operation"),
    ("后端scraper_id", "catalogOpId"),
    ("测试输入", "input"),
    ("耗时(ms)", "timeMs"),
    ("返回条数", "records"),
    ("task_id", "taskId"),
    ("错误/现象", "error"),
    ("故障归属", "domain"),
    ("严重级", "severity"),
    ("给后端的建议", "advice"),
]

# Column widths tuned per field type (short enum-y columns narrow, free-text wide).
COLUMN_WIDTHS: dict[str, int] = {
    "状态": 10,
    "工具": 26,
    "平台/目标": 20,
    "operation": 22,
    "后端scraper_id": 26,
    "测试输入": 46,
    "耗时(ms)": 10,
    "返回条数": 10,
    "task_id": 30,
    "错误/现象": 60,
    "故障归属": 14,
    "严重级": 8,
    "给后端的建议": 50,
}

SEVERITY_RANK = {"P0": 0, "P1": 1, "P2": 2, "P3": 3}


def find_latest_report() -> str:
    candidates = sorted(glob.glob(os.path.join(REPORTS_DIR, "full-*.json")))
    if not candidates:
        raise FileNotFoundError(
            f"No monitoring/reports/full-*.json found in {REPORTS_DIR} — "
            "run monitoring/smoke/full-tools-probe.mjs first."
        )
    # Filenames are ISO-timestamp-sortable (see full-tools-probe.mjs's
    # isoForFilename), so a plain sort gives us the newest last.
    return candidates[-1]


def severity_sort_key(row: dict[str, Any]) -> tuple[int, str]:
    """Worst -> best: P0 first, then P1/P2/P3, passing rows (no severity) last."""
    sev = row.get("severity")
    rank = SEVERITY_RANK.get(sev, len(SEVERITY_RANK))
    return (rank, str(row.get("name", "")))


def cell_value(row: dict[str, Any], field: str) -> Any:
    v = row.get(field)
    if v is None:
        return "-"
    return v


def build_results_sheet(ws: Worksheet, results: list[dict[str, Any]]) -> None:
    headers = [h for h, _ in COLUMNS]
    ws.append(headers)
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")

    sorted_results = sorted(results, key=severity_sort_key)
    status_col_idx = 1  # "状态" is always the first column per COLUMNS above

    for row_data in sorted_results:
        row_values = [cell_value(row_data, field) for _, field in COLUMNS]
        ws.append(row_values)
        row_idx = ws.max_row
        status = str(row_data.get("status", ""))
        fill = STATUS_FILLS.get(status)
        if fill is not None:
            status_cell = ws.cell(row=row_idx, column=status_col_idx)
            status_cell.fill = fill
            font = STATUS_FONTS.get(status)
            if font is not None:
                status_cell.font = font

    # Column widths
    for col_idx, (header, _) in enumerate(COLUMNS, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = COLUMN_WIDTHS.get(header, 18)

    # Autofilter over the full used range, frozen header row.
    last_col_letter = get_column_letter(len(COLUMNS))
    ws.auto_filter.ref = f"A1:{last_col_letter}{ws.max_row}"
    ws.freeze_panes = "A2"


def build_summary_sheet(ws: Worksheet, report: dict[str, Any]) -> None:
    summary = report.get("summary", {}) or {}
    headers = ["指标", "值"]
    ws.append(headers)
    for col_idx in (1, 2):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Headline metric (code review fix, 2026-07-24): `maxSeverity` mixes
    # domains — a P0 among our own tools (①/②) and a P0 among backend-only
    # (③) findings look identical there. `maxOursSeverity` is scoped to ①/②
    # rows only, so this sheet can never misread a fully backend-only event
    # as "we got paged". It is listed FIRST and highlighted below; the mixed
    # `maxSeverity` is kept as secondary context right after it.
    headline_label = "我方最高严重级 (maxOursSeverity) — 是否该我们背锅，headline"
    rows: list[tuple[str, Any]] = [
        (headline_label, summary.get("maxOursSeverity") or "-"),
        ("全局最高严重级 (maxSeverity，含后端③)", summary.get("maxSeverity") or "-"),
        ("开始时间 (UTC)", report.get("startedAt") or "-"),
        ("结束时间 (UTC)", report.get("finishedAt") or "-"),
        ("MCP_URL", report.get("mcpUrl") or "-"),
        ("在线工具数 (tools/list)", report.get("liveToolCount", "-")),
        ("我方问题数 (①/②)", summary.get("oursCount", 0)),
        ("后端问题数 (③)", summary.get("backendCount", 0)),
    ]

    by_status = summary.get("byStatus", {}) or {}
    for status_name in ("PASS", "SLOW", "FAIL", "MISSING"):
        if status_name in by_status:
            rows.append((f"状态计数 - {status_name}", by_status[status_name]))
    for status_name, count in by_status.items():
        if status_name not in ("PASS", "SLOW", "FAIL", "MISSING"):
            rows.append((f"状态计数 - {status_name}", count))

    by_severity = summary.get("bySeverity", {}) or {}
    for sev_name in ("P0", "P1", "P2", "P3"):
        if sev_name in by_severity:
            rows.append((f"严重级计数 - {sev_name}", by_severity[sev_name]))

    missing_tools = summary.get("missingTools", []) or []
    rows.append(("缺失工具 (missingTools)", ", ".join(missing_tools) if missing_tools else "-"))

    fatal_error = summary.get("fatalError")
    if fatal_error:
        rows.append(("FATAL (endpoint down?)", fatal_error))

    rows.append(("exit code", report.get("exitCode", "-")))

    # Light tint distinguishing "our own severity" from the "backend-inclusive
    # (informational) severity" row directly below it — same brand-purple
    # family as the header fill, at low opacity.
    headline_fill = PatternFill(start_color="EDE7FA", end_color="EDE7FA", fill_type="solid")
    headline_font = Font(bold=True)

    for label, value in rows:
        ws.append([label, value])
        if label == headline_label:
            row_idx = ws.max_row
            for col_idx in (1, 2):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.fill = headline_fill
                cell.font = headline_font

    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 70
    ws.freeze_panes = "A2"


def write_csv(csv_path: str, results: list[dict[str, Any]]) -> None:
    sorted_results = sorted(results, key=severity_sort_key)
    headers = [h for h, _ in COLUMNS]
    # utf-8-sig writes the UTF-8 BOM so Excel on Windows/macOS opens Chinese
    # headers correctly instead of mangling them as Latin-1/mojibake.
    with open(csv_path, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.writer(f)
        writer.writerow(headers)
        for row_data in sorted_results:
            writer.writerow([cell_value(row_data, field) for _, field in COLUMNS])


def main() -> int:
    if len(sys.argv) > 1:
        json_path = sys.argv[1]
    else:
        json_path = find_latest_report()

    print(f"[render-report] reading {json_path}")
    with open(json_path, "r", encoding="utf-8") as f:
        report = json.load(f)

    results = report.get("results", []) or []
    stem = os.path.splitext(os.path.basename(json_path))[0]  # e.g. "full-2026-07-24T02-17-00-000Z"

    wb = Workbook()
    results_ws = wb.active
    results_ws.title = "逐工具测试结果"
    build_results_sheet(results_ws, results)

    summary_ws = wb.create_sheet("汇总")
    build_summary_sheet(summary_ws, report)

    os.makedirs(REPORTS_DIR, exist_ok=True)
    xlsx_path = os.path.join(REPORTS_DIR, f"{stem}.xlsx")
    csv_path = os.path.join(REPORTS_DIR, f"{stem}.csv")

    wb.save(xlsx_path)
    write_csv(csv_path, results)

    print(f"[render-report] wrote {xlsx_path}")
    print(f"[render-report] wrote {csv_path}")
    print(f"[render-report] {len(results)} row(s), maxSeverity={report.get('summary', {}).get('maxSeverity')}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
